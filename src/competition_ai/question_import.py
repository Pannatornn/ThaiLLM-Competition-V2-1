from __future__ import annotations

import base64
import csv
import io
from dataclasses import dataclass
from typing import Iterable

from openpyxl import load_workbook


MAX_UPLOAD_BYTES = 8 * 1024 * 1024
MAX_QUESTIONS = 500
QUESTION_HEADERS = {
    "question",
    "questions",
    "query",
    "prompt",
    "คำถาม",
    "โจทย์",
}


@dataclass(frozen=True)
class ImportedQuestions:
    filename: str
    sheet: str | None
    header: str
    questions: list[str]
    row_numbers: list[int]


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\ufeff", "").strip()


def _normalize_header(value) -> str:
    return " ".join(_clean(value).casefold().split())


def _find_question_column(rows: list[list[object]]) -> tuple[int, int, str]:
    """Return (header_row_index, column_index, display_header)."""
    for ridx, row in enumerate(rows[:20]):
        for cidx, value in enumerate(row):
            normalized = _normalize_header(value)
            if normalized in QUESTION_HEADERS:
                return ridx, cidx, _clean(value) or "question"

    # Fallback for organizer-style blank templates that contain questions but
    # use an unexpected header: choose the first text-heavy column.
    if rows:
        width = max((len(r) for r in rows[:30]), default=0)
        best_col = -1
        best_score = -1
        for cidx in range(width):
            score = 0
            for row in rows[:30]:
                value = _clean(row[cidx] if cidx < len(row) else "")
                if len(value) >= 8:
                    score += 1
            if score > best_score:
                best_col = cidx
                best_score = score
        if best_col >= 0 and best_score > 0:
            return 0, best_col, "question"

    raise ValueError(
        "ไม่พบคอลัมน์คำถาม กรุณาใช้หัวคอลัมน์ question, query, prompt, คำถาม หรือ โจทย์"
    )


def _extract(rows: list[list[object]], filename: str, sheet: str | None) -> ImportedQuestions:
    if not rows:
        raise ValueError("ไฟล์ไม่มีข้อมูล")

    header_row, question_col, header = _find_question_column(rows)
    questions: list[str] = []
    row_numbers: list[int] = []

    start = header_row + 1
    for ridx in range(start, len(rows)):
        row = rows[ridx]
        value = _clean(row[question_col] if question_col < len(row) else "")
        if not value:
            continue
        questions.append(value)
        row_numbers.append(ridx + 1)
        if len(questions) >= MAX_QUESTIONS:
            break

    if not questions:
        # If the fallback selected row 1 as data rather than a true header,
        # include it when it looks like a question.
        if header == "question" and rows:
            first = _clean(rows[0][question_col] if question_col < len(rows[0]) else "")
            if first and _normalize_header(first) not in QUESTION_HEADERS:
                questions.append(first)
                row_numbers.append(1)

    if not questions:
        raise ValueError("ไม่พบคำถามที่ไม่ว่างในไฟล์")

    return ImportedQuestions(
        filename=filename,
        sheet=sheet,
        header=header,
        questions=questions,
        row_numbers=row_numbers,
    )


def parse_csv_bytes(data: bytes, filename: str) -> ImportedQuestions:
    text = None
    last_error = None
    for encoding in ("utf-8-sig", "utf-8", "cp874", "tis-620"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError as exc:
            last_error = exc
    if text is None:
        raise ValueError(f"อ่าน encoding ของ CSV ไม่ได้: {last_error}")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    rows = [list(row) for row in csv.reader(io.StringIO(text), dialect)]
    return _extract(rows, filename, None)


def parse_xlsx_bytes(data: bytes, filename: str) -> ImportedQuestions:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        # Search every sheet so organizer files do not need a fixed sheet name.
        errors: list[str] = []
        for ws in wb.worksheets:
            rows = [list(row) for row in ws.iter_rows(values_only=True)]
            try:
                return _extract(rows, filename, ws.title)
            except ValueError as exc:
                errors.append(f"{ws.title}: {exc}")
        raise ValueError("ไม่พบตารางคำถามใน workbook: " + " | ".join(errors[:4]))
    finally:
        wb.close()


def decode_upload(filename: str, content_base64: str) -> ImportedQuestions:
    name = _clean(filename)
    if not name:
        raise ValueError("filename is required")

    try:
        data = base64.b64decode(content_base64, validate=True)
    except Exception as exc:
        raise ValueError("ไฟล์อัปโหลดไม่ใช่ Base64 ที่ถูกต้อง") from exc

    if not data:
        raise ValueError("ไฟล์ว่าง")
    if len(data) > MAX_UPLOAD_BYTES:
        raise ValueError("ไฟล์ใหญ่เกิน 8 MB")

    lower = name.casefold()
    if lower.endswith(".csv"):
        return parse_csv_bytes(data, name)
    if lower.endswith(".xlsx"):
        return parse_xlsx_bytes(data, name)

    raise ValueError("รองรับเฉพาะไฟล์ .xlsx และ .csv")
